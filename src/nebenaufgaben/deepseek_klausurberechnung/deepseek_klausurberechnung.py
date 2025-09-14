import os
import json
import re
import requests
import pandas as pd
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import sympy as sp

OLLAMA_URL = "http://localhost:11434/api/generate"

def get_model_response(prompt):
    response = requests.post(
        OLLAMA_URL,
        json={"model": "deepseek-r1:32b", "prompt": prompt, "stream": False}
    )
    if response.status_code == 200:
        return response.json().get("response", "(Keine Antwort erhalten)").strip()
    else:
        return f"Fehler {response.status_code}: {response.text}"

model = SentenceTransformer("paraphrase-MiniLM-L6-v2")

def semantische_aehnlichkeit(text1, text2):
    emb1 = model.encode(text1, convert_to_tensor=True)
    emb2 = model.encode(text2, convert_to_tensor=True)
    return util.cos_sim(emb1, emb2).item()

def enthaelt_schluesselwoerter(antwort, basetruth):
    antwort = antwort.lower()
    return all(wort in antwort for wort in basetruth.lower().split())

# Unicode Maps für Tief- und Hochstellung (kleine Zahlen/Buchstaben)
subscript_map = {
    '0':'₀', '1':'₁', '2':'₂', '3':'₃', '4':'₄',
    '5':'₅', '6':'₆', '7':'₇', '8':'₈', '9':'₉',
    'a':'ₐ', 'e':'ₑ', 'h':'ₕ', 'i':'ᵢ', 'j':'ⱼ',
    'k':'ₖ', 'l':'ₗ', 'm':'ₘ', 'n':'ₙ', 'o':'ₒ',
    'p':'ₚ', 'r':'ᵣ', 's':'ₛ', 't':'ₜ', 'u':'ᵤ',
    'v':'ᵥ', 'x':'ₓ'
}

superscript_map = {
    '0':'⁰', '1':'¹', '2':'²', '3':'³', '4':'⁴',
    '5':'⁵', '6':'⁶', '7':'⁷', '8':'⁸', '9':'⁹',
    'a':'ᵃ', 'b':'ᵇ', 'c':'ᶜ', 'd':'ᵈ', 'e':'ᵉ',
    'f':'ᶠ', 'g':'ᵍ', 'h':'ʰ', 'i':'ⁱ', 'j':'ʲ',
    'k':'ᵏ', 'l':'ˡ', 'm':'ᵐ', 'n':'ⁿ', 'o':'ᵒ',
    'p':'ᵖ', 'r':'ʳ', 's':'ˢ', 't':'ᵗ', 'u':'ᵘ',
    'v':'ᵛ', 'w':'ʷ', 'x':'ˣ', 'y':'ʸ', 'z':'ᶻ'
}

def to_subscript(s):
    return ''.join(subscript_map.get(ch, ch) for ch in s)

def to_superscript(s):
    return ''.join(superscript_map.get(ch, ch) for ch in s)

def latex_to_plaintext(text):
    # Spezialfall: \int_0^x → ∫₀ˣ
    def integral_replacer(match):
        sub = to_subscript(match.group(1))
        sup = to_superscript(match.group(2))
        return f"∫{sub}{sup}"

    text = re.sub(r"\\int_([\w]+)\^([\w]+)", integral_replacer, text)

    ersetzungen = [
        (r"\\int", "∫"),
        (r"\\sqrt\{([^{}]+)\}", r"√(\1)"),
        (r"\\frac\{([^{}]+)\}\{([^{}]+)\}", r"(\1)/(\2)"),
        (r"\\pm", "±"),
        (r"\\cdot", "·"),
        (r"\\infty", "∞"),
        (r"\\left\(|\\right\(", "("),
        (r"\\left\)|\\right\)", ")"),
        (r"\\begin\{[^}]+\}|\\end\{[^}]+\}", ""),
        (r"\\text\{([^{}]*)\}", r"\1"),
        (r"\\[a-zA-Z]+", ""),
        (r"\^\{([^{}]+)\}", r"^\1"),
        (r"\_\{([^{}]+)\}", r"_\1"),
        (r"\\", "")
    ]
    for pattern, repl in ersetzungen:
        text = re.sub(pattern, repl, text)
    return text.strip()

def extrahiere_final_answer_robust(text):
    text = re.sub(r'<think>.*?</think>', '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r'\\boxed\{([^{}]*)\}', r'\1', text)
    text = re.sub(r'\\boxed\s*\n*\\begin\{aligned\}.*?\\end\{aligned\}', '', text, flags=re.DOTALL)
    marker_regex = r'(?:\*\*?\s*)?(Antwort:|Final Answer|Endg\u00fcltige Antwort)(?:\s*\*\*?)?:?\s*'
    match = re.search(marker_regex, text, flags=re.IGNORECASE)
    if match:
        start = match.end()
        text = text[start:]
    text = latex_to_plaintext(text)
    return text.strip()

def pruefe_mathematische_gleichheit(text1, text2):
    x = sp.symbols('x')
    y = sp.Function('y')(x)
    def parse_eq(eq_text):
        eq_text = eq_text.replace("y'", "Derivative(y, x)")
        eq_text = re.sub(r"y\\s*\\(\\s*x\\s*\\)", "y", eq_text)
        if '=' in eq_text:
            lhs, rhs = eq_text.split('=', 1)
        else:
            lhs, rhs = eq_text, "0"
        lhs_expr = sp.sympify(lhs.strip(), locals={'Derivative': sp.Derivative, 'y': y, 'x': x})
        rhs_expr = sp.sympify(rhs.strip(), locals={'Derivative': sp.Derivative, 'y': y, 'x': x})
        return lhs_expr, rhs_expr
    try:
        lhs1, rhs1 = parse_eq(text1)
        lhs2, rhs2 = parse_eq(text2)
        return sp.simplify((lhs1 - rhs1) - (lhs2 - rhs2)) == 0
    except:
        return False

def bewerte_antwort(antwort, basetruth):
    final_answer = extrahiere_final_answer_robust(antwort)
    if pruefe_mathematische_gleichheit(final_answer, basetruth):
        return "✔️ korrekt", 1.0, final_answer
    score = semantische_aehnlichkeit(final_answer, basetruth)
    if score > 0.85 or enthaelt_schluesselwoerter(final_answer, basetruth):
        bewertung = "✔️ korrekt"
    elif score > 0.6:
        bewertung = "➖ teilkorrekt"
    else:
        bewertung = "❌ falsch"
    return bewertung, round(score, 3), final_answer

alle_dateien = input("Möchtest du alle JSON-Dateien im Verzeichnis verarbeiten? (j/n): ").strip().lower()

if alle_dateien == 'j':
    zu_verarbeiten = [f for f in os.listdir() if f.endswith(".json")]
else:
    print("Verfügbare JSON-Dateien:")
    alle_jsons = [f for f in os.listdir() if f.endswith(".json")]
    for i, fname in enumerate(alle_jsons):
        print(f"{i+1}. {fname}")
    eingabe = input("Bitte gib die Nummern ein (kommagetrennt, z.B. 1,2,3): ").strip()
    nummern = [int(num) - 1 for num in eingabe.split(',') if num.strip().isdigit() and 0 < int(num) <= len(alle_jsons)]
    zu_verarbeiten = [alle_jsons[i] for i in nummern]

for filename in zu_verarbeiten:
    print(f"Verarbeite Datei: {filename}")
    with open(filename, "r", encoding="utf-8") as f:
        aufgaben = json.load(f)

    daten = []
    for aufgabe in aufgaben:
        frage = aufgabe["frage"]
        basetruth = aufgabe["basetruth"]
        print(f"Frage (ID {aufgabe['id']}): {frage}")
        antwort = get_model_response(frage).replace('\\', '')
        print(f"Antwort DeepSeek (roh):\n{antwort}")
        bewertung, score, final_answer = bewerte_antwort(antwort, basetruth)
        print(f"Antwort extrahiert: {final_answer}")
        print(f"Ähnlichkeit: {score:.3f} → Bewertung: {bewertung}")
        daten.append({
            "ID": aufgabe["id"],
            "Frage": frage,
            "Basetruth": basetruth,
            "Antwort DeepSeek (roh)": antwort,
            "Antwort DeepSeek (final)": final_answer,
            "Ähnlichkeit": score,
            "Bewertung": bewertung
        })
        print("="*40)

    excel_name = os.path.splitext(filename)[0] + ".xlsx"
    df = pd.DataFrame(daten)
    df.to_excel(excel_name, index=False)
    print(f"Gespeichert: {excel_name}")

    wb = load_workbook(excel_name)
    ws = wb.active

    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    MAX_BREITE = 100
    
    headers = [cell.value for cell in ws[1]]
    for idx, col in enumerate(ws.columns):
        header = headers[idx]
        if header == "Antwort DeepSeek (roh)":
            continue
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        max_len = min(max_len, MAX_BREITE)
        col_letter = get_column_letter(idx+1)
        ws.column_dimensions[col_letter].width = max_len + 2

    farbe_map = {
        "✔️ korrekt": "C6EFCE",
        "❌ falsch": "FFC7CE",
        "➖ teilkorrekt": "FFEB9C"
    }
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for cell in row:
            farbe = farbe_map.get(cell.value)
            if farbe:
                cell.fill = PatternFill(start_color=farbe, end_color=farbe, fill_type="solid")

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.000'

    wb.save(excel_name)
    print(f"Excel-Datei formatiert und gespeichert: {excel_name}")

